import sys
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.Qt import QTableWidgetItem
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QTableWidget,

)

from PyQt5 import QtCore, QtGui, QtWidgets


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1126, 694)
        MainWindow.setStyleSheet("background-color: rgb(233, 255, 157);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(30, 30, 441, 31))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(30, 100, 251, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(30, 240, 201, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(30, 390, 171, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(510, 100, 301, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(510, 240, 377, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(510, 390, 201, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.listView = QtWidgets.QListView(self.centralwidget)
        self.listView.setGeometry(QtCore.QRect(30, 140, 331, 81))
        self.listView.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listView.setObjectName("listView")
        self.listView_2 = QtWidgets.QListView(self.centralwidget)
        self.listView_2.setGeometry(QtCore.QRect(30, 280, 331, 81))
        self.listView_2.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listView_2.setObjectName("listView_2")
        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(35, 421, 331, 91))
        self.listWidget.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listWidget.setObjectName("listWidget")
        self.listWidget_2 = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_2.setGeometry(QtCore.QRect(510, 140, 331, 81))
        self.listWidget_2.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listWidget_2.setObjectName("listWidget_2")
        self.listWidget_3 = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_3.setGeometry(QtCore.QRect(510, 280, 331, 81))
        self.listWidget_3.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listWidget_3.setObjectName("listWidget_3")
        self.listWidget_4 = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_4.setGeometry(QtCore.QRect(510, 430, 331, 81))
        self.listWidget_4.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.listWidget_4.setObjectName("listWidget_4")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(750, 570, 311, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.pushButton.setObjectName("pushButton")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1126, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Анализ безопасности системы Умный дом"))
        self.label.setText(_translate("MainWindow", "Выберите параметры Умного дома"))
        self.label_2.setText(_translate("MainWindow", "Умные световые решения:"))
        self.label_3.setText(_translate("MainWindow", "Умная бытовая техника:"))
        self.label_4.setText(_translate("MainWindow", "Модули управления:"))
        self.label_5.setText(_translate("MainWindow", "Умные камеры видеонаблюдения:"))
        self.label_6.setText(_translate("MainWindow", "Умные устройства открытия гаражных ворот:"))
        self.label_7.setText(_translate("MainWindow", "Датчики Умного дома:"))
        self.pushButton.setText(_translate("MainWindow", "Анализировать"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

def appication():
    app=QApplication(sys.argv)
    window = QMainWindow()

    window.setWindowTitle("Smart home")
    window.setGeometry(300, 250, 300, 200)

    window.show()
    sys.exit(app.exec_())

if __name__=="__main__":
    appication()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setMinimumWidth(1200)
        self.setMinimumHeight(600)

        layout = QVBoxLayout()

        self.table = QTableWidget(self)
        self.table.setRowCount(4)
        self.table.setColumnCount(4)
        layout.addWidget(self.table)

        btn = QPushButton("Загрузить")
        btn.clicked.connect(self.btn_click)
        layout.addWidget(btn)

        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

    def btn_click(self):
        # Import `load_workbook` module from `openpyxl`
        from openpyxl import load_workbook

        # Load in the workbook
        wb = load_workbook('./123.xlsx')

        # Get sheet names
        sheet = wb['Sheet1']
        print(sheet.cell(row=2, column=1).value)
        for row in range(1, 5):
            for column in range(1, 5):
                item = QTableWidgetItem()
                item.setText(str(sheet.cell(row=row, column=column).value))
                self.table.setItem(row-1, column-1, item)


app = QApplication(sys.argv)

window = MainWindow()
window.show()

app.exec()
